import csv
import multiprocessing
from openpyxl import load_workbook

STOP_CODON = "***STOP***"

queues = {}
processes = []

def csvWriter(filename, queue):
	with open(filename + ".csv", "wb") as csvFile:
		writer = csv.writer(
			csvFile,
			lineterminator="\n",
			delimiter=",",
			quoting=csv.QUOTE_NONNUMERIC
		)

		while True:
			line = queue.get()

			if line == STOP_CODON:
				return

			print "WRITING ({0})".format(str(line))
			writer.writerow(line)

def parseWorksheet(ws):
	queue = queues[ws.title]

	for row in ws.rows:
		data = []

		for cell in row:
			print "READING ({0})".format(str(cell.value))
			data.append(cell.value)

		queue.put(data)
	queue.put(STOP_CODON)

def parseWorkbook(wb):
	for ws in wb.worksheets:
		queue = multiprocessing.Queue()
		writerProcess = multiprocessing.Process(target=csvWriter, args=[ws.title, queue])
		writerProcess.start()
		processes.append(writerProcess)
		queues[ws.title] = queue

		parseWorksheet(ws)

def main():
	parseWorkbook(load_workbook(filename="input.xlsx", read_only=True, data_only=True))

	for process in processes:
		process.join()

if __name__ == "__main__":
	main()